from openpyxl import Workbook, load_workbook
import random

wb = Workbook()
dest_filename = 'random_square.xlsx'
for j in range(5):
    wb.create_sheet(title = 'Sheet_' + str(j))
sh_rd = wb[wb.sheetnames[random.randint(0, 5)]]
print(sh_rd)
col = random.randint(1, 10)
rw = random.randint(1, 10)
for i in range(col, col + 12):
    for k in range(rw, rw + 12):
        sh_rd.cell(column = i, row = k, value = random.random())
wb.save(filename = dest_filename)