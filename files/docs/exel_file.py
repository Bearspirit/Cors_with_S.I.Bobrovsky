from openpyxl import Workbook, load_workbook

wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws2 = wb.create_sheet(title='Test')
ws2['F5'] = 3.14
ws2.cell(column = 5, row = 4, value = '12345')
wb.save(filename = dest_filename)
wb = load_workbook(filename = dest_filename)
ws = wb['Test']
print(ws['F5'].value)